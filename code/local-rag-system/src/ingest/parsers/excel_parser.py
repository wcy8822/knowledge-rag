from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from pathlib import Path

from .base_parser import BaseParser
from ...models.document import DocumentWithChunks, DocumentChunkWithMetadata, ExcelMetadata, DataSource

class ExcelParser(BaseParser):
    """Excel文件解析器"""
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.primary_key_patterns = self.config.get('primary_key_patterns', 
            ['id', 'ID', 'key', 'KEY', '编号', 'ID号'])
        self.max_rows_per_chunk = self.config.get('max_rows_per_chunk', 50)
        self.skip_empty_rows = self.config.get('skip_empty_rows', True)
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否为Excel文件"""
        return Path(file_path).suffix.lower() in ['.xlsx', '.xls']
    
    def parse(self, file_path: str) -> DocumentWithChunks:
        """解析Excel文件"""
        doc_id = self._generate_doc_id(file_path)
        file_size = self._get_file_size(file_path)
        file_hash = self._calculate_file_hash(file_path)
        
        # 创建文档对象
        document = DocumentWithChunks(
            doc_id=doc_id,
            file_path=file_path,
            file_type=DataSource.EXCEL,
            file_size_bytes=file_size,
            sha256_hash=file_hash
        )
        
        try:
            # 读取Excel文件
            chunks = self._parse_excel_sheets(file_path, doc_id)
            document.chunks = chunks
            document.processing_status = "completed"
            
        except Exception as e:
            document.processing_status = "failed"
            print(f"Error parsing Excel file {file_path}: {e}")
            raise
        
        return document
    
    def _parse_excel_sheets(self, file_path: str, doc_id: str) -> List[DocumentChunkWithMetadata]:
        """解析Excel工作表"""
        chunks = []
        
        try:
            # 使用openpyxl读取工作表
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet_chunks = self._parse_sheet(wb[sheet_name], sheet_name, doc_id)
                chunks.extend(sheet_chunks)
                
        except Exception as e:
            # 回退到pandas
            try:
                excel_data = pd.read_excel(file_path, sheet_name=None)
                for sheet_name, df in excel_data.items():
                    sheet_chunks = self._parse_dataframe(df, sheet_name, doc_id)
                    chunks.extend(sheet_chunks)
            except Exception as e2:
                print(f"Failed to parse Excel with both methods: {e2}")
                raise
        
        return chunks
    
    def _parse_sheet(self, sheet, sheet_name: str, doc_id: str) -> List[DocumentChunkWithMetadata]:
        """解析单个工作表"""
        chunks = []
        chunk_index = 0
        
        # 获取表头
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value or f"Column_{len(headers)+1}")
        
        # 查找主键
        primary_key = self._find_primary_key(headers)
        
        # 处理数据行
        current_chunk_rows = []
        row_start = 2  # Excel行号从1开始，数据从第2行开始
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if self.skip_empty_rows and not any(cell is not None for cell in row):
                continue
            
            current_chunk_rows.append((row_num, row))
            
            # 达到块大小或工作表结束时创建块
            if len(current_chunk_rows) >= self.max_rows_per_chunk:
                chunk = self._create_chunk_from_rows(
                    current_chunk_rows, headers, sheet_name, 
                    primary_key, doc_id, chunk_index
                )
                chunks.append(chunk)
                chunk_index += 1
                current_chunk_rows = []
                row_start = row_num + 1
        
        # 处理剩余行
        if current_chunk_rows:
            chunk = self._create_chunk_from_rows(
                current_chunk_rows, headers, sheet_name,
                primary_key, doc_id, chunk_index
            )
            chunks.append(chunk)
        
        return chunks
    
    def _parse_dataframe(self, df: pd.DataFrame, sheet_name: str, doc_id: str) -> List[DocumentChunkWithMetadata]:
        """使用DataFrame解析工作表"""
        chunks = []
        chunk_index = 0
        
        # 清理列名
        df = df.copy()
        df.columns = df.columns.astype(str)
        headers = list(df.columns)
        
        # 查找主键
        primary_key = self._find_primary_key(headers)
        
        # 分块处理
        total_rows = len(df)
        for start_idx in range(0, total_rows, self.max_rows_per_chunk):
            end_idx = min(start_idx + self.max_rows_per_chunk, total_rows)
            chunk_df = df.iloc[start_idx:end_idx]
            
            # 跳过空行
            if self.skip_empty_rows:
                chunk_df = chunk_df.dropna(how='all')
            
            if chunk_df.empty:
                continue
            
            # 创建块
            chunk = self._create_chunk_from_dataframe(
                chunk_df, headers, sheet_name, primary_key,
                doc_id, chunk_index, start_idx + 2  # +2 对应Excel行号
            )
            chunks.append(chunk)
            chunk_index += 1
        
        return chunks
    
    def _create_chunk_from_rows(self, rows_data: List[tuple], headers: List[str], 
                               sheet_name: str, primary_key: Optional[str], 
                               doc_id: str, chunk_index: int) -> DocumentChunkWithMetadata:
        """从行数据创建文档块"""
        chunk_contents = []
        row_start = None
        row_end = None
        primary_key_value = None
        
        for row_num, row in rows_data:
            if row_start is None:
                row_start = row_num
            row_end = row_num
            
            # 创建行描述
            row_desc_parts = [f"表格【{sheet_name}】第{row_num}行："]
            for col_idx, (header, cell_value) in enumerate(zip(headers, row)):
                if cell_value is not None and str(cell_value).strip():
                    row_desc_parts.append(f"{header}={cell_value}")
                    
                    # 记录主键值
                    if header == primary_key and primary_key_value is None:
                        primary_key_value = str(cell_value)
            
            chunk_contents.append("，".join(row_desc_parts))
        
        # 创建块内容
        content = "\n".join(chunk_contents)
        
        # 创建Excel元数据
        excel_metadata = ExcelMetadata(
            sheet_name=sheet_name,
            row_start=row_start or 0,
            row_end=row_end or 0,
            primary_key=primary_key_value,
            column_count=len(headers)
        )
        
        # 创建文档块
        chunk = DocumentChunkWithMetadata(
            doc_id=doc_id,
            content=content,
            chunk_index=chunk_index,
            token_count=len(content.split()),  # 简单的token计数
            metadata={
                "sheet_name": sheet_name,
                "row_start": row_start,
                "row_end": row_end,
                "primary_key": primary_key_value,
                "column_count": len(headers)
            },
            excel_metadata=excel_metadata
        )
        
        return chunk
    
    def _create_chunk_from_dataframe(self, df: pd.DataFrame, headers: List[str],
                                   sheet_name: str, primary_key: Optional[str],
                                   doc_id: str, chunk_index: int, start_row: int) -> DocumentChunkWithMetadata:
        """从DataFrame创建文档块"""
        chunk_contents = []
        
        for idx, (_, row) in enumerate(df.iterrows()):
            row_num = start_row + idx
            row_desc_parts = [f"表格【{sheet_name}】第{row_num}行："]
            
            primary_key_value = None
            for header, cell_value in row.items():
                if pd.notna(cell_value) and str(cell_value).strip():
                    row_desc_parts.append(f"{header}={cell_value}")
                    
                    # 记录主键值
                    if header == primary_key and primary_key_value is None:
                        primary_key_value = str(cell_value)
            
            chunk_contents.append("，".join(row_desc_parts))
        
        # 创建块内容
        content = "\n".join(chunk_contents)
        
        # 创建Excel元数据
        end_row = start_row + len(df) - 1
        excel_metadata = ExcelMetadata(
            sheet_name=sheet_name,
            row_start=start_row,
            row_end=end_row,
            primary_key=primary_key_value,
            column_count=len(headers)
        )
        
        # 创建文档块
        chunk = DocumentChunkWithMetadata(
            doc_id=doc_id,
            content=content,
            chunk_index=chunk_index,
            token_count=len(content.split()),
            metadata={
                "sheet_name": sheet_name,
                "row_start": start_row,
                "row_end": end_row,
                "primary_key": primary_key_value,
                "column_count": len(headers)
            },
            excel_metadata=excel_metadata
        )
        
        return chunk
    
    def _find_primary_key(self, headers: List[str]) -> Optional[str]:
        """在表头中查找主键"""
        for header in headers:
            if header and any(pattern.lower() in str(header).lower() 
                           for pattern in self.primary_key_patterns):
                return header
        return None
    
    def _generate_doc_id(self, file_path: str) -> str:
        """生成文档ID"""
        return f"excel_{Path(file_path).stem}"